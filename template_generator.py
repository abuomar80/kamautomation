"""
Excel Template Generator for Advanced Configuration
This module creates a downloadable Excel template with all required sheets and columns
"""
import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl.worksheet.datavalidation import DataValidation

def generate_excel_template():
    """
    Generates an Excel template file with all required sheets and sample data
    Returns: BytesIO object containing the Excel file
    """
    # Create an Excel writer object
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # Sheet 1: Material_Types
        material_types_df = pd.DataFrame({
            'Legacy System': ['Example Legacy Type 1', 'Example Legacy Type 2'],
            'Description': ['Description for type 1', 'Description for type 2'],
            'Medad': ['Book', 'Journal']
        })
        material_types_df.to_excel(writer, sheet_name='Material_Types', index=False)
        
        # Sheet 2: Statistical_Codes
        statistical_codes_df = pd.DataFrame({
            'Medad Statistical Type': ['Collection', 'Branch'],
            'Medad Statistical Code': ['REF', 'MAIN']
        })
        statistical_codes_df.to_excel(writer, sheet_name='Statistical_Codes', index=False)
        
        # Sheet 3: Item_status
        # Note: Item_status sheet may not be actively used, but it's required
        item_status_df = pd.DataFrame({
            'Status': ['Available', 'Checked out', 'On order'],
            'Code': ['Available', 'Checked out', 'On order']
        })
        item_status_df.to_excel(writer, sheet_name='Item_status', index=False)
        
        # Sheet 4: User_groups
        # User groups only need: Legacy System (optional), Description (optional, defaults to Medad), Medad (required - Patron group name)
        # Departments are separate and loaded from Department sheet
        user_groups_df = pd.DataFrame({
            'Legacy System': ['Example Legacy Group 1', 'Example Legacy Group 2', 'Example Legacy Group 3'],
            'Description': ['Faculty members', 'Students', 'Staff members'],
            'Medad': ['Faculty', 'Student', 'Staff']
        })
        user_groups_df.to_excel(writer, sheet_name='User_groups', index=False)
        
        # Sheet 5: Location
        location_df = pd.DataFrame({
            'ServicePoints name': ['Main Library Desk', 'Reference Desk'],
            'ServicePoints Codes': ['MLD', 'REF'],
            'InstitutionsName': ['Main Institution', 'Main Institution'],
            'InstitutionsCodes': ['MI', 'MI'],
            'CampusNames': ['Main Campus', 'Main Campus'],
            'CampusCodes': ['MC', 'MC'],
            'LibrariesName': ['Main Library', 'Main Library'],
            'LibrariesCodes': ['ML', 'ML'],
            'LocationsName': ['Main Library Desk', 'Reference Desk'],
            'LocationsCodes': ['MLD', 'REF']
        })
        location_df.to_excel(writer, sheet_name='Location', index=False)
        
        # Sheet 6: Calendar
        calendar_df = pd.DataFrame({
            'ServicePoints name': ['Main Library Desk'],
            'start': ['09:00'],
            'end': ['17:00']
        })
        calendar_df.to_excel(writer, sheet_name='Calendar', index=False)
        
        # Sheet 7: Calendar Exceptions
        calendar_exceptions_df = pd.DataFrame({
            'ServicePoints name': ['Main Library Desk'],
            'name': ['New Year Holiday'],
            'startDate': ['2024-01-01'],
            'endDate': ['2024-01-01'],
            'allDay': [True]
        })
        calendar_exceptions_df.to_excel(writer, sheet_name='Calendar Exceptions', index=False)
        
        # Sheet 8: Department (optional - departments can also be defined in User_groups sheet)
        # This sheet is kept for backward compatibility but departments in User_groups will be created automatically
        department_df = pd.DataFrame({
            'Name': ['Main Department', 'Reference Department'],
            'Code': ['MAIN', 'REF']
        })
        department_df.to_excel(writer, sheet_name='Department', index=False)
        
        # Sheet 9: FeeFineOwner
        # Service Points should be entered as labels (names), not UUIDs
        # Use "ALL" to assign to all service points
        fee_fine_owner_df = pd.DataFrame({
            'Service Points': ['Main Library Desk', 'ALL'],
            'Owner': ['1', '2']
        })
        fee_fine_owner_df.to_excel(writer, sheet_name='FeeFineOwner', index=False)
        
        # Sheet 10: FeeFine
        # owner should reference the owner value from FeeFineOwner sheet
        # feeFineType: type identifier (string)
        # amountWithoutVat: numeric amount
        # vat: VAT rate (optional, numeric)
        # automatic: boolean (true/false)
        # id: optional UUID (leave empty to auto-generate)
        fee_fine_df = pd.DataFrame({
            'feeFineType': ['1', '2'],
            'amountWithoutVat': [5.0, 10.0],
            'vat': [0.05, 0.10],
            'owner': ['1', '2'],
            'automatic': [False, True],
            'id': ['', '']  # Leave empty to auto-generate
        })
        fee_fine_df.to_excel(writer, sheet_name='FeeFine', index=False)
        
        # Sheet 11: Waives
        # nameReason: reason for waiving (required)
        # id: optional UUID (leave empty to auto-generate)
        # Each row creates one waive record
        waives_df = pd.DataFrame({
            'nameReason': ['Manager Approval', 'Lost Item Found'],
            'id': ['', '']  # Leave empty to auto-generate
        })
        waives_df.to_excel(writer, sheet_name='Waives', index=False)
        
        # Sheet 12: PaymentMethods
        # nameMethod: payment method name (required)
        # allowedRefundMethod: boolean (optional, default: false)
        # owner: references Owner from FeeFineOwner sheet (required)
        # id: optional UUID (leave empty to auto-generate)
        # Each row creates one payment method record
        payment_methods_df = pd.DataFrame({
            'nameMethod': ['cash', 'credit card'],
            'allowedRefundMethod': [False, True],
            'owner': ['1', '2'],
            'id': ['', '']  # Leave empty to auto-generate
        })
        payment_methods_df.to_excel(writer, sheet_name='PaymentMethods', index=False)
        
        # Sheet 13: Refunds
        # nameReason: reason for refund (required)
        # id: optional UUID (leave empty to auto-generate)
        # Each row creates one refund record
        refunds_df = pd.DataFrame({
            'nameReason': ['error', 'duplicate payment'],
            'id': ['', '']  # Leave empty to auto-generate
        })
        refunds_df.to_excel(writer, sheet_name='Refunds', index=False)
        
        # Sheet 14: LoanPolicies
        # Complex structure with nested objects for loansPolicy and renewalsPolicy
        # name: policy name (required)
        # description: policy description (optional)
        # loanable: boolean (true/false, default: true)
        # renewable: boolean (true/false, default: true)
        # profileId: loan profile ID (e.g., "Rolling", default: "Rolling")
        # periodDuration: loan period duration in days/weeks/months (numeric)
        # periodIntervalId: period interval ("Days", "Weeks", "Months")
        # closedLibraryDueDateManagementId: closed library due date management (e.g., "END_OF_THE_NEXT_OPEN_DAY")
        # gracePeriodDuration: grace period duration (numeric)
        # gracePeriodIntervalId: grace period interval ("Days", "Weeks", "Months")
        # itemLimit: maximum number of items (numeric, as string)
        # numberAllowed: number of renewals allowed (numeric, as string)
        # renewFromId: renewal from date ("CURRENT_DUE_DATE" or "SYSTEM_DATE")
        # id: optional UUID (leave empty to auto-generate)
        loan_policies_df = pd.DataFrame({
            'name': ['Standard Loan Policy', 'Short Term Loan Policy'],
            'description': ['Standard loan policy for regular items', 'Short term loan policy for high-demand items'],
            'loanable': [True, True],
            'renewable': [True, True],
            'unlimitedRenewals': [False, False],  # New column for unlimited renewals
            'profileId': ['Rolling', 'Rolling'],
            'periodDuration': [15, 7],
            'periodIntervalId': ['Days', 'Days'],
            'closedLibraryDueDateManagementId': ['END_OF_THE_NEXT_OPEN_DAY', 'END_OF_THE_NEXT_OPEN_DAY'],
            'gracePeriodDuration': [3, 1],
            'gracePeriodIntervalId': ['Days', 'Days'],
            'itemLimit': ['5', '3'],
            'numberAllowed': ['3', '1'],  # Required only if unlimitedRenewals is False
            'renewFromId': ['CURRENT_DUE_DATE', 'CURRENT_DUE_DATE'],  # Required only if unlimitedRenewals is False
            'id': ['', '']  # Leave empty to auto-generate
        })
        loan_policies_df.to_excel(writer, sheet_name='LoanPolicies', index=False)
        
        # Access the workbook to add data validation (while still in context manager)
        workbook = writer.book
        worksheet = workbook['LoanPolicies']
        
        # Define dropdown options based on the screenshot
        interval_options = "Minutes,Hours,Days,Weeks,Months"
        renew_from_options = "CURRENT_DUE_DATE,SYSTEM_DATE"
        closed_library_options = "KEEP_CURRENT_DATE,END_OF_THE_PREVIOUS_OPEN_DAY,END_OF_THE_NEXT_OPEN_DAY"
        boolean_options = "TRUE,FALSE"
        
        # Find column indices
        period_interval_col = None
        grace_period_interval_col = None
        renew_from_col = None
        closed_library_col = None
        renewable_col = None
        unlimited_renewals_col = None
        number_allowed_col = None
        
        for idx, col in enumerate(loan_policies_df.columns, start=1):
            if col == 'periodIntervalId':
                period_interval_col = idx
            elif col == 'gracePeriodIntervalId':
                grace_period_interval_col = idx
            elif col == 'renewFromId':
                renew_from_col = idx
            elif col == 'closedLibraryDueDateManagementId':
                closed_library_col = idx
            elif col == 'renewable':
                renewable_col = idx
            elif col == 'unlimitedRenewals':
                unlimited_renewals_col = idx
            elif col == 'numberAllowed':
                number_allowed_col = idx
        
        # Add data validation for periodIntervalId
        if period_interval_col:
            # Apply to all rows (starting from row 2, excluding header) and extend to row 1000 for future entries
            max_row = max(len(loan_policies_df) + 1, 1000)
            col_letter = worksheet.cell(2, period_interval_col).column_letter
            period_interval_dv = DataValidation(type="list", formula1=f'"{interval_options}"', allow_blank=True)
            period_interval_dv.add(f"{col_letter}2:{col_letter}{max_row}")
            worksheet.add_data_validation(period_interval_dv)
        
        # Add data validation for gracePeriodIntervalId
        if grace_period_interval_col:
            max_row = max(len(loan_policies_df) + 1, 1000)
            col_letter = worksheet.cell(2, grace_period_interval_col).column_letter
            grace_period_interval_dv = DataValidation(type="list", formula1=f'"{interval_options}"', allow_blank=True)
            grace_period_interval_dv.add(f"{col_letter}2:{col_letter}{max_row}")
            worksheet.add_data_validation(grace_period_interval_dv)
        
        # Add data validation for renewFromId
        if renew_from_col:
            max_row = max(len(loan_policies_df) + 1, 1000)
            col_letter = worksheet.cell(2, renew_from_col).column_letter
            renew_from_dv = DataValidation(type="list", formula1=f'"{renew_from_options}"', allow_blank=True)
            renew_from_dv.add(f"{col_letter}2:{col_letter}{max_row}")
            worksheet.add_data_validation(renew_from_dv)
        
        # Add data validation for closedLibraryDueDateManagementId
        if closed_library_col:
            max_row = max(len(loan_policies_df) + 1, 1000)
            col_letter = worksheet.cell(2, closed_library_col).column_letter
            closed_library_dv = DataValidation(type="list", formula1=f'"{closed_library_options}"', allow_blank=True)
            closed_library_dv.add(f"{col_letter}2:{col_letter}{max_row}")
            worksheet.add_data_validation(closed_library_dv)
        
        # Add data validation for renewable (true/false)
        if renewable_col:
            max_row = max(len(loan_policies_df) + 1, 1000)
            col_letter = worksheet.cell(2, renewable_col).column_letter
            renewable_dv = DataValidation(type="list", formula1=f'"{boolean_options}"', allow_blank=True)
            renewable_dv.add(f"{col_letter}2:{col_letter}{max_row}")
            worksheet.add_data_validation(renewable_dv)
        
        # Add data validation for unlimitedRenewals (true/false)
        if unlimited_renewals_col:
            max_row = max(len(loan_policies_df) + 1, 1000)
            col_letter = worksheet.cell(2, unlimited_renewals_col).column_letter
            unlimited_renewals_dv = DataValidation(type="list", formula1=f'"{boolean_options}"', allow_blank=True)
            unlimited_renewals_dv.add(f"{col_letter}2:{col_letter}{max_row}")
            worksheet.add_data_validation(unlimited_renewals_dv)
        
        # Add conditional validation for numberAllowed and renewFromId
        # These should be optional if unlimitedRenewals is TRUE
        # Note: Excel doesn't support dynamic validation based on other cells directly,
        # but we'll allow blank values and handle validation in the Python code
        if number_allowed_col:
            max_row = max(len(loan_policies_df) + 1, 1000)
            col_letter = worksheet.cell(2, number_allowed_col).column_letter
            # Allow numeric values or blank (blank if unlimitedRenewals is TRUE)
            number_allowed_dv = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
            number_allowed_dv.add(f"{col_letter}2:{col_letter}{max_row}")
            worksheet.add_data_validation(number_allowed_dv)
        
        if renew_from_col:
            max_row = max(len(loan_policies_df) + 1, 1000)
            col_letter = worksheet.cell(2, renew_from_col).column_letter
            # Allow dropdown values or blank (blank if unlimitedRenewals is TRUE)
            renew_from_dv = DataValidation(type="list", formula1=f'"{renew_from_options}"', allow_blank=True)
            renew_from_dv.add(f"{col_letter}2:{col_letter}{max_row}")
            worksheet.add_data_validation(renew_from_dv)
    
    output.seek(0)
    return output

def download_template_button():
    """
    Creates a download button in Streamlit for the Excel template
    Generates a fresh template with all 14 sheets including:
    Material_Types, Statistical_Codes, Item_status, User_groups, Location, 
    Calendar, Calendar Exceptions, Department, FeeFineOwner, FeeFine, 
    Waives, PaymentMethods, Refunds, LoanPolicies
    
    Note: The template includes data validation dropdowns for:
    - renewable (TRUE, FALSE)
    - unlimitedRenewals (TRUE, FALSE) - if TRUE, numberAllowed and renewFromId are optional
    - periodIntervalId (Minutes, Hours, Days, Weeks, Months)
    - gracePeriodIntervalId (Minutes, Hours, Days, Weeks, Months)
    - renewFromId (CURRENT_DUE_DATE, SYSTEM_DATE) - optional if unlimitedRenewals is TRUE
    - closedLibraryDueDateManagementId (KEEP_CURRENT_DATE, END_OF_THE_PREVIOUS_OPEN_DAY, END_OF_THE_NEXT_OPEN_DAY)
    - numberAllowed (numeric, >= 0) - optional if unlimitedRenewals is TRUE
    """
    import datetime
    import time
    import re
    
    # Get tenant name from session state
    tenant_name = st.session_state.get('tenant') or st.session_state.get('tenant_name', '')
    
    # Clean tenant name for filename (remove invalid characters)
    if tenant_name:
        # Replace invalid filename characters with underscore
        tenant_name = re.sub(r'[<>:"/\\|?*]', '_', str(tenant_name))
        tenant_name = tenant_name.strip()
        # Add prefix with separator
        tenant_prefix = f"{tenant_name}____"
    else:
        tenant_prefix = ""
    
    # Generate fresh template each time (no caching)
    # Force regeneration by adding a small delay to ensure unique timestamp
    time.sleep(0.01)  # Small delay to ensure unique timestamp
    template_file = generate_excel_template()
    
    # Get the bytes data - this ensures we're not passing a cached BytesIO object
    template_bytes = template_file.getvalue()
    
    # Add timestamp with microseconds to filename to help with browser caching issues
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    
    # Use a unique key that changes each time to prevent Streamlit caching
    unique_key = f"download_template_{timestamp}_{time.time()}"
    
    # Build filename with tenant prefix
    file_name = f"{tenant_prefix}Advanced_Configuration_Template_{timestamp}.xlsx"
    
    st.download_button(
        label="📥 Download Excel Template (14 Sheets with Data Validation)",
        data=template_bytes,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Download the Excel template with all 14 required sheets including data validation dropdowns for LoanPolicies fields",
        key=unique_key  # Unique key to prevent caching
    )
